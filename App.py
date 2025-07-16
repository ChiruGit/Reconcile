###############
#### PILOT. #####
###############

import streamlit as st
import pandas as pd
import numpy as np
import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import requests



st.set_page_config(page_title="Chiran Reconciliation Web App", layout="wide")
st.title("Chiran's Reconciliation Web App")

st.markdown("""
Upload your **First** and **Second Excel** files below to perform reconciliation.  

""")

col1, col2 = st.columns(2)
with col1:
    file1 = st.file_uploader("Upload File1 Excel", type=["xlsx"], key='file1')
with col2:
    file2 = st.file_uploader("Upload File2 Excel", type=["xlsx"], key='file2')

# Read into DataFrames

if all([file1, file2]):
    # Load the Aspire and Safaricom files
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)
    

# Identify exact matches by Transaction ID, Description, and Amount
    matches = pd.merge(df1, df2, on=["Transaction ID", "Description", "Amount"])

# Find entries only in Sheet1 (based on Transaction ID)
    only_in_df1 = df1[~df1['Transaction ID'].isin(df2['Transaction ID'])]

# Find entries only in Sheet2 (based on Transaction ID)
    only_in_df2 = df2[~df2['Transaction ID'].isin(df1['Transaction ID'])]

# --- 5. Prepare workbook ---
    output = io.BytesIO()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Matched"
    for r in matches.itertuples(index=False):
        ws1.append(list(r))
    ws1.insert_rows(1)
    ws1.insert_rows(1)
    for i, c in enumerate(list(matches.columns), 1):
        ws1.cell(row=2, column=i, value=c)
    #style_header(ws1, header_row=2)
    #autofit(ws1)

    ws2 = wb.create_sheet("Only_in_firstsheet")
    for r in only_in_df1.itertuples(index=False):
        ws2.append(list(r))
    ws2.insert_rows(1)
    ws2.insert_rows(1)
    for i, c in enumerate(list(only_in_df1.columns), 1):
        ws2.cell(row=2, column=i, value=c)
    #style_header(ws2, header_row=2)
    #autofit(ws2)

    ws3 = wb.create_sheet("Only_in_Secondsheet")
    for r in only_in_df2.itertuples(index=False):
        ws3.append(list(r))
    ws3.insert_rows(1)
    ws3.insert_rows(1)
    for i, c in enumerate(list(only_in_df2.columns), 1):
        ws3.cell(row=2, column=i, value=c)
    #style_header(ws3, header_row=2)
    #autofit(ws3)

   
    wb.save(output)
    output.seek(0)

    st.success("Excel workbook with all reports is ready.")
    st.download_button(
        label="Download All Reports (Excel workbook)",
        data=output,
        file_name="ChiAwesome_Reconciliation_Reports.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.info("Sheets: Matches, Only_in_firstsheet, Only_in_Secondsheet")
else:
    st.info("Please upload both File1 and File2 Excel files to proceed.")
